"""Data-driven execution — feed rows into ``${var}`` scripts.

Runtime variables plus ``AC_for_each`` already let a script iterate over a
list, but the list had to be hard-coded. This module supplies the missing
*data source*: load tabular rows from CSV, JSON, SQLite, Excel, or an
inline literal, then drive the same action body once per row.

A source is a plain dict so it round-trips through JSON action files,
the socket server, and the REST API::

    {"kind": "csv",    "path": "users.csv"}
    {"kind": "json",   "path": "cases.json"}
    {"kind": "sqlite", "path": "app.db", "query": "SELECT * FROM users"}
    {"kind": "excel",  "path": "matrix.xlsx", "sheet": "Sheet1"}
    {"kind": "inline", "rows": [{"user": "a"}, {"user": "b"}]}

:func:`load_rows` always returns ``List[Dict[str, Any]]`` so each row can
be bound to a single variable whose keys are addressable as
``${row.user}`` downstream.
"""
from __future__ import annotations

import csv
import json
import os
from contextlib import closing
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional
from urllib.parse import quote

from je_auto_control.utils.sqlite_support import require_sqlite3

_READ_ONLY_SQL_PREFIXES = ("select", "with")


def _resolve_path(raw: str) -> Path:
    """Resolve a user-supplied path and verify it points at a real file."""
    if not raw or not isinstance(raw, str):
        raise ValueError("data source 'path' must be a non-empty string")
    resolved = Path(os.path.realpath(os.path.expanduser(raw)))
    if not resolved.is_file():
        raise FileNotFoundError(f"data source file not found: {resolved}")
    return resolved


def _load_csv(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Load rows from a CSV file; the header row supplies the dict keys."""
    path = _resolve_path(source["path"])
    delimiter = str(source.get("delimiter", ","))
    encoding = str(source.get("encoding", "utf-8"))
    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        return [dict(row) for row in reader]


def _coerce_json_rows(payload: Any) -> List[Dict[str, Any]]:
    """Normalise parsed JSON into a list of row dicts."""
    if isinstance(payload, dict):
        payload = payload.get("rows", [])
    if not isinstance(payload, list):
        raise ValueError("JSON data source must be a list or {'rows': [...]}")
    rows: List[Dict[str, Any]] = []
    for item in payload:
        if not isinstance(item, dict):
            raise ValueError("each JSON data row must be an object")
        rows.append(dict(item))
    return rows


def _load_json(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Load rows from a JSON file (a list of objects, or ``{"rows": [...]}``)."""
    path = _resolve_path(source["path"])
    encoding = str(source.get("encoding", "utf-8"))
    with path.open("r", encoding=encoding) as handle:
        return _coerce_json_rows(json.load(handle))


def _validate_select(query: str) -> str:
    """Reject anything but a single read-only SELECT/WITH statement."""
    cleaned = query.strip().rstrip(";").strip()
    if ";" in cleaned:
        raise ValueError("sqlite data source allows a single statement only")
    if not cleaned.lower().startswith(_READ_ONLY_SQL_PREFIXES):
        raise ValueError("sqlite data source query must start with SELECT/WITH")
    return cleaned


def _load_sqlite(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Run a read-only SELECT against a SQLite file and return dict rows."""
    path = _resolve_path(source["path"])
    query = _validate_select(str(source["query"]))
    # SQLite treats ``?`` as the query separator and percent-decodes ``%XX`` in
    # the path, so a raw f-string opens the wrong file when the path contains
    # ``?``/``#``/``%``. Percent-encode those while keeping the separators and
    # the Windows drive colon literal.
    safe_path = quote(str(path), safe="/:\\")
    uri = f"file:{safe_path}?mode=ro"
    # closing(): the sqlite3 context manager commits/rolls back but never closes
    # the connection, so the file handle leaks until GC. Close it explicitly.
    driver = require_sqlite3()
    with closing(driver.connect(uri, uri=True)) as conn:
        conn.row_factory = driver.Row
        rows = conn.execute(query).fetchall()
    return [dict(row) for row in rows]


def _load_excel(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Load rows from an .xlsx file; the first row supplies the dict keys."""
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Excel data sources require openpyxl (pip install openpyxl).",
        ) from error
    path = _resolve_path(source["path"])
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_name = source.get("sheet")
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return []
        keys = [str(cell) for cell in header]
        return [dict(zip(keys, values)) for values in rows_iter]
    finally:
        workbook.close()


def _load_inline(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return rows supplied directly in the spec under ``rows``."""
    return _coerce_json_rows(source.get("rows", []))


_LOADERS: Dict[str, Callable[[Dict[str, Any]], List[Dict[str, Any]]]] = {
    "csv": _load_csv,
    "json": _load_json,
    "sqlite": _load_sqlite,
    "excel": _load_excel,
    "inline": _load_inline,
}


def supported_kinds() -> List[str]:
    """Return the data-source kinds this module can load."""
    return sorted(_LOADERS)


def load_rows(source: Dict[str, Any],
              limit: Optional[int] = None) -> List[Dict[str, Any]]:
    """Load tabular rows from ``source``; return a list of row dicts.

    :param source: a spec dict with a ``kind`` key (one of
        :func:`supported_kinds`) plus kind-specific fields.
    :param limit: optional cap on the number of rows returned.
    """
    if not isinstance(source, dict):
        raise ValueError("data source must be a dict with a 'kind' key")
    kind = str(source.get("kind", "")).strip().lower()
    loader = _LOADERS.get(kind)
    if loader is None:
        raise ValueError(
            f"unknown data source kind {kind!r}; "
            f"expected one of {supported_kinds()}"
        )
    rows = loader(source)
    if limit is not None and limit >= 0:
        rows = rows[: int(limit)]
    return rows
